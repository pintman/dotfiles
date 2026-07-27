#!/usr/bin/env python3
"""Erzeugt die Preisrecherche-Excel-Tabelle aus einer JSON-Konfiguration.

Nimmt recherchierte Rohdaten (Bezeichnung/Preis brutto/Link je Artikel und
Haendler) entgegen und uebernimmt die komplette Formatierung sowie die
Netto-Berechnung (Brutto / 1.19) -- das muss nicht mehr pro Lauf aus dem
SKILL.md nachgebaut werden.
"""
import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HAENDLER_FARBEN = {
    "Conrad": "FFD700",
    "Reichelt": "4472C4",
    "Alternate": "ED7D31",
    "AZ Delivery": "70AD47",
    "Berrybase": "9B59B6",
    "Voelkner": "E74C3C",
}
FARBEN_FALLBACK = ["8E8E8E", "5B9BD5", "A9D18E", "FFC000", "C00000", "7030A0"]

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(name="Arial", size=10)
LINK_FONT = Font(name="Arial", size=10, color="0563C1", underline="single")
GUENSTIG_FONT = Font(name="Arial", size=10, bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
DUENN = Side(style="thin", color="BFBFBF")
RAHMEN = Border(left=DUENN, right=DUENN, top=DUENN, bottom=DUENN)
GRAU = PatternFill("solid", fgColor="F2F2F2")
ARTIKEL_KOPF_FARBE = "404040"
GUENSTIG_FARBE = "70AD47"

PREIS_FORMAT = '#.##0,00 "€"'
NICHT_GEFUNDEN = "Nicht gefunden"
SUB_HEADER = ["Artikelbezeichnung", "Preis netto (€)", "Preis brutto (€)", "Link"]

MWST_SATZ = 1.19


def farbe_fuer(haendler: str, index: int) -> str:
    return HAENDLER_FARBEN.get(haendler, FARBEN_FALLBACK[index % len(FARBEN_FALLBACK)])


def netto(brutto):
    if brutto is None:
        return None
    return round(brutto / MWST_SATZ, 2)


def style_zelle(cell, font, align=WRAP, fill=None):
    cell.font = font
    cell.alignment = align
    cell.border = RAHMEN
    if fill is not None:
        cell.fill = fill


def build(config: dict):
    haendler_liste = config["haendler"]
    artikel_liste = config["artikel"]
    if not haendler_liste:
        raise ValueError("Konfiguration enthaelt keine Haendler")
    if not artikel_liste:
        raise ValueError("Konfiguration enthaelt keine Artikel")

    wb = Workbook()
    ws = wb.active
    ws.title = "Preisrecherche"

    spalte = 2  # B
    haendler_spalten = {}
    for h in haendler_liste:
        haendler_spalten[h] = spalte
        spalte += 4
    guenstig_start = spalte

    # Kopfzeile: Spalte A (Artikel), vertikal ueber beide Kopfzeilen verbunden
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    c = ws.cell(row=1, column=1, value="Gesuchter Artikel")
    style_zelle(c, HEADER_FONT, HEADER_ALIGN, PatternFill("solid", fgColor=ARTIKEL_KOPF_FARBE))
    style_zelle(ws.cell(row=2, column=1), HEADER_FONT, HEADER_ALIGN, PatternFill("solid", fgColor=ARTIKEL_KOPF_FARBE))

    # Kopfzeile: je Haendler ein Gruppenheader (Zeile 1, ueber 4 Spalten) + Subheader (Zeile 2)
    for idx, h in enumerate(haendler_liste):
        col = haendler_spalten[h]
        fill = PatternFill("solid", fgColor=farbe_fuer(h, idx))
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 3)
        style_zelle(ws.cell(row=1, column=col, value=h), HEADER_FONT, HEADER_ALIGN, fill)
        for j, sub in enumerate(SUB_HEADER):
            style_zelle(ws.cell(row=2, column=col + j, value=sub), HEADER_FONT, HEADER_ALIGN, fill)

    # Kopfzeile: "Guenstigster Anbieter" ueber die letzten zwei Spalten
    fill_guenstig = PatternFill("solid", fgColor=GUENSTIG_FARBE)
    ws.merge_cells(start_row=1, start_column=guenstig_start, end_row=1, end_column=guenstig_start + 1)
    style_zelle(ws.cell(row=1, column=guenstig_start, value="Günstigster Anbieter (netto)"), HEADER_FONT, HEADER_ALIGN, fill_guenstig)
    for j, sub in enumerate(["Anbieter", "Preis netto (€)"]):
        style_zelle(ws.cell(row=2, column=guenstig_start + j, value=sub), HEADER_FONT, HEADER_ALIGN, fill_guenstig)

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30

    ws.column_dimensions[get_column_letter(1)].width = 35
    for h in haendler_liste:
        col = haendler_spalten[h]
        ws.column_dimensions[get_column_letter(col)].width = 30
        ws.column_dimensions[get_column_letter(col + 1)].width = 14
        ws.column_dimensions[get_column_letter(col + 2)].width = 14
        ws.column_dimensions[get_column_letter(col + 3)].width = 12
    ws.column_dimensions[get_column_letter(guenstig_start)].width = 20
    ws.column_dimensions[get_column_letter(guenstig_start + 1)].width = 14

    nicht_gefunden_zaehler = 0
    zeile = 3
    for artikel_idx, artikel in enumerate(artikel_liste):
        ws.row_dimensions[zeile].height = 40
        schattiert = artikel_idx % 2 == 1
        fill = GRAU if schattiert else None

        style_zelle(ws.cell(row=zeile, column=1, value=artikel["name"]), DATA_FONT, WRAP, fill)

        netto_zellen = []
        for h in haendler_liste:
            col = haendler_spalten[h]
            angebot = artikel.get("angebote", {}).get(h) or {}
            bezeichnung = angebot.get("bezeichnung")
            brutto = angebot.get("preis_brutto")
            link = angebot.get("link")

            if bezeichnung is None and brutto is None:
                nicht_gefunden_zaehler += 1

            netto_wert = netto(brutto)
            werte = [
                bezeichnung if bezeichnung else NICHT_GEFUNDEN,
                netto_wert if netto_wert is not None else NICHT_GEFUNDEN,
                brutto if brutto is not None else NICHT_GEFUNDEN,
                "Link" if link else NICHT_GEFUNDEN,
            ]

            zellen = []
            for j, wert in enumerate(werte):
                cc = ws.cell(row=zeile, column=col + j, value=wert)
                style_zelle(cc, DATA_FONT, WRAP, fill)
                if j in (1, 2) and isinstance(wert, (int, float)):
                    cc.number_format = PREIS_FORMAT
                zellen.append(cc)

            if link:
                zellen[3].hyperlink = link
                zellen[3].font = LINK_FONT

            if netto_wert is not None:
                netto_zellen.append(get_column_letter(col + 1) + str(zeile))

        if netto_zellen:
            bereich = ",".join(netto_zellen)
            min_formel = f"=MIN({bereich})"
            namen_array = "{" + ",".join(f'"{h}"' for h in haendler_liste) + "}"
            anbieter_formel = f"=INDEX({namen_array},MATCH(MIN({bereich}),({bereich}),0))"
        else:
            min_formel = NICHT_GEFUNDEN
            anbieter_formel = NICHT_GEFUNDEN

        c1 = ws.cell(row=zeile, column=guenstig_start, value=anbieter_formel)
        c2 = ws.cell(row=zeile, column=guenstig_start + 1, value=min_formel)
        style_zelle(c1, GUENSTIG_FONT, WRAP, fill)
        style_zelle(c2, GUENSTIG_FONT, WRAP, fill)
        if min_formel != NICHT_GEFUNDEN:
            c2.number_format = PREIS_FORMAT

        zeile += 1

    ws.freeze_panes = "B3"
    return wb, len(artikel_liste), haendler_liste, nicht_gefunden_zaehler


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", required=True, type=Path, help="Pfad zur JSON-Konfiguration")
    parser.add_argument("--output", type=Path, help="Ueberschreibt das 'output'-Feld aus der Config")
    args = parser.parse_args()

    try:
        config = json.loads(args.config.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        print(f"Fehler beim Lesen der Konfiguration: {exc}", file=sys.stderr)
        sys.exit(1)

    try:
        wb, anzahl_artikel, haendler_liste, nicht_gefunden = build(config)
    except (KeyError, ValueError) as exc:
        print(f"Fehler in der Konfiguration: {exc}", file=sys.stderr)
        sys.exit(1)

    output = args.output or Path(config["output"])
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    print(f"Gespeichert: {output}")
    print(f"Artikel: {anzahl_artikel}")
    print(f"Haendler: {', '.join(haendler_liste)}")
    print(f"Nicht gefunden: {nicht_gefunden}")


if __name__ == "__main__":
    main()
